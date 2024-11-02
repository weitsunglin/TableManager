# -*- coding: utf-8 -*-

import openpyxl
import json
import io
import ast
import os
import xml.etree.ElementTree as ET
import time
import shutil

CONFIG_FILE_NAME = "config.json"
TABLES_FILE_NAME = "TableSetting.ts"

# 將 Unicode 整數陣列轉成一般整數陣列，範例：u'[205,18364]' >> [205,18364]
def parse_unicode_list_to_list(s):
    if isinstance(s, str) and not s.startswith(u"['"):
        try:
            v = ast.literal_eval(s)
            if isinstance(v, list):
                return v
            else:
                raise ValueError("ValueError")
        except (SyntaxError, ValueError):
            pass
    return s

# 解析 Excel 文件，並轉換為 JSON 格式輸出
def parse_excel(excel_file, json_f_name, index):
    jd = []  # 儲存 JSON 內容的列表
    heads = []  # 儲存表頭
    book = openpyxl.load_workbook(excel_file)
    sheet = book.active

    max_row = sheet.max_row
    max_column = sheet.max_column

    # 讀取表頭
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)

    # 讀取表格資料並處理每一行
    for row in range(max_row):
        if row < 2:
            continue

        one_line = {}
        for column in range(max_column):
            k = heads[column]
            v = sheet.cell(row + 1, column + 1).value
            if v is not None:
                v = parse_unicode_list_to_list(v)
                one_line[k] = v

        jd.append(one_line)

    book.close()
    jd = [item for item in jd if any(item.values())]  # 移除空白列

    config = load_config()
    generate_extends(json_f_name, config["out_put_extends_path"], json_f_name, jd)
    generate_tableSettings(json_f_name, index)
    generate_json_file(jd, config["out_put_table_path"], json_f_name)

# 將資料寫入 JSON 文件
def generate_json_file(jd, json_file_path, json_f_name):
    json_file_path = os.path.join(json_file_path, f"{json_f_name}.json")
    if jd:
        with io.open(json_file_path, 'w', encoding='utf-8') as f:
            txt = json.dumps(jd, indent=2, ensure_ascii=False)
            txt = fix_json_string(txt)
            f.write(txt)
            print("成功輸出 JSON 檔案")

# 修正 JSON 字串格式
def fix_json_string(json_str):
    json_str = json_str.replace('"[', '[')  # 替換 "[ 為 [
    json_str = json_str.replace(']"', ']')  # 替換 ]" 為 ]
    json_str = json_str.replace("'", '"')   # 替換單引號為雙引號
    json_str = json_str.replace('\\"', '')  # 移除轉義符號
    return json_str

# 產生 TypeScript 檔案的表格設定
def generate_tableSettings(json_f_name, index):
    json_f_name_array.insert(index, json_f_name)

    if index < excel_files_count:
        return

    typescript_file_path = os.path.join("../../tableManager", TABLES_FILE_NAME)

    with io.open(typescript_file_path, 'w', encoding='utf-8') as ts_file:
        ts_file.write("export enum Tables {\n")
        for enum_name in json_f_name_array:
            ts_file.write(f"    {enum_name} = '{enum_name}',\n")
        ts_file.write("}")

    print(f"成功寫入 TypeScript 檔案：{typescript_file_path}")

# 列出資料夾中的所有 Excel 檔案
def list_excel_files(folder_path):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    return excel_files

# 載入設定檔
def load_config() -> dict:
    with open(CONFIG_FILE_NAME, "r") as config_file:
        return json.load(config_file)

# 獲取 Excel 路徑和 JSON 路徑
def get_paths(config: dict, excel_file_name: str) -> tuple:
    excel_file_path = os.path.join(config["excel_folder_path"], excel_file_name)
    return excel_file_path

# 產生 TypeScript 擴展檔案
def generate_extends(file_name, output_dir, json_f_name, jd) -> None:
    class_name = f"{file_name}Extend"
    little_class_name = lowercase_string(class_name)

    ts_code = "export enum ID {\n    // 自行新增 ID KeyValue\n}\n\n"
    ts_code += "export enum ColumnName {\n"

    # 產生 Enum 成員
    if jd and isinstance(jd[0], dict):
        keys = jd[0].keys()
        for key in keys:
            ts_code += f"    {key} = '{key}',\n"

    ts_code += "}\n\n"
    ts_code += f"class {class_name} {{\n    constructor() {{\n    }}\n}}\n\n"
    ts_code += f"const {little_class_name} = new {class_name}();\n"
    ts_code += f"export default {little_class_name};"

    output_file_path = os.path.join(output_dir, f"{class_name}.ts")

    if os.path.exists(output_file_path):
        print(f"檔案已存在：{output_file_path}")
        return

    with open(output_file_path, 'w', encoding='utf-8') as ts_file:
        ts_file.write(ts_code)
    print(f"成功寫入檔案：{output_file_path}")

# 將字串轉換為小寫
def lowercase_string(s):
    if not s:
        return s
    return s.lower()

# 清除 JSON 快取檔案
def cleanJsonCache():
    if os.path.exists(config["out_put_table_path"]):
        for filename in os.listdir(config["out_put_table_path"]):
            file_path = os.path.join(config["out_put_table_path"], filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f'刪除失敗 {file_path}。原因: {e}')
        print("清除 JSON 快取。")
    else:
        print("目錄不存在。")

if __name__ == "__main__":
    start_time = time.time()
  
    json_f_name_array = []

    try:
        config = load_config()

        cleanJsonCache()

        # 列出指定資料夾中的所有 Excel 檔案
        excel_files = list_excel_files(config["excel_folder_path"])
        print("找到的 Excel 檔案:", excel_files)

        excel_files_count = len(excel_files)

        if not excel_files:
            print("未在 excels 資料夾中找到 Excel 檔案。")
        else:
            # 處理每個 Excel 檔案
            for i, excel_file_name in enumerate(excel_files, start=1):
                file_name_without_brackets = os.path.splitext(excel_file_name)[0]
                excel_file_path = get_paths(config, excel_file_name)

                try:
                    parse_excel(excel_file_path, file_name_without_brackets, i)
                except Exception as e:
                    print(f"處理檔案 {excel_file_name} 時出錯：{e}")

        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"腳本執行時間：{elapsed_time:.2f} 秒。")
    except Exception as e:
        print(f"腳本錯誤：{e}")
